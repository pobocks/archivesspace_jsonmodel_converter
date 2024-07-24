# Archival objects
from asnake.jsonmodel import JM
from psycopg.rows import dict_row
import openpyxl
from openpyxl.styles import NamedStyle, Protection, Font, Border, Side
from textwrap import dedent

# Dynamic globals, set via config
client = None
xw = None
conn = None
log = None

# Static globals
tablename = 'tblitems'
resource_tablename = 'tblcolls'
dept_id = 48

# Each AO needs:
#   to be associated with its resource on creation
#   that resource to have its link added to its instances
#   TODO to be associated with its accession on creation
#   TODO that accession to be added to its instances

def construct_title(row):
    output = f"{row['itemname'] or 'Unnamed'} - "
    output += row['title'] or 'Untitled'
    return output

# Required AO Fields:
#  THEORETICALLY only title, resource, and level
def process_archival_objects(tablename, resource_tablename, dept_id, hooks=None):
    if not hooks:
        hooks = []
    with conn.cursor(row_factory=dict_row) as cur:
        # get a count
        cur.execute(f'''SELECT COUNT(*) FROM "{tablename}"
                        WHERE deptcodeid = {dept_id}''')
        count = cur.fetchone()['count']
        log.info(f"Processing {count} archival objects in table {tablename}")
        for row in cur.execute(
                f'''SELECT collid,
                           itemid,
                           itemname,
                           title,
                           itemdesc
                    FROM {tablename}
                    WHERE collid IS NOT null
                      AND itemid !~* '^[L|R]'
                      AND deptcodeid = {dept_id}
                 ORDER BY collid ASC, datecreated ASC'''):
            aid = xw.get_aspace_id(resource_tablename, row["collid"])
            if aid is None:
                log.error(f"Resource with original collid of {row['collid']} missing from crosswalk for item {row['itemid']}, skipping")
                continue
            record_json = JM.archival_object(
                component_id=row['itemid'],
                title=construct_title(row),
                level="item",
                resource={"ref":aid},
                external_ids=[
                    JM.external_id(external_id="COLLID: " + str(row['collid']), source="access"),
                    JM.external_id(external_id=str(row['itemid']), source="access")
                ],
                publish=True)
            if row['itemdesc']:
                record_json['notes'] = [
                    JM.note_singlepart(
                        type="abstract",
                        label="Abstract from itemdesc",
                        content=[
                            row['itemdesc']
                        ],
                        publish=True
                    )
                ]
            # Code to alter the archival object or link it to other records goes here, or
            # can be passed in the hooks argument, in which case it should have the following signature:
            #   def hook(tablename, resource_tablename, dept_id, json, row) -> dict
            # and make global references as needed, and return the altered record_json
            for hook in hooks:
                record_json = hook(tablename, resource_tablename, dept_id, record_json, row)
            yield row['collid'], row['itemid'], record_json

def setup_globals(config, input_log, which={'aspace', 'xw', 'conn', 'log'}):
    global client, xw, conn, log
    if 'log' in which:
        log = input_log
    if 'aspace' in which:
        client = config['d']['aspace']
        client.authorize()
    if 'xw' in which:
        xw = config['d']['crosswalk']
        xw.create_crosswalk()
    if 'conn' in which:
        conn = config['d']['postgres']

good_statii = frozenset(('Created', 'Updated',))
def archival_objects_create(config, input_log):
    setup_globals(config, input_log)
    for collid, orig_id, json in process_archival_objects(tablename, resource_tablename, dept_id):
        ao_id = xw.get_aspace_id(tablename, orig_id)
        if ao_id is not None:
            ao = client.get(ao_id).json()
            json['lock_version'] = ao['lock_version']
            res = client.post(ao_id, json=json).json()
        else:
            res = client.post('repositories/2/archival_objects', json=json).json()
        if 'status' in res and res['status'] in good_statii:
            aspace_uri = res['uri']
            xw.add_or_update(tablename, orig_id, 'archival_object', aspace_uri)
            log.info(f'Added or updated {tablename} {orig_id}')
        elif 'error' in res:
            error = res['error']
            if 'conflicting_record' in error:
                aspace_id = error['conflicting_record'][0]
                log.warn(f'Possible duplicate for ID {orig_id} in {tablename} found with URI of {aspace_id}')
            if 'source' in error:
                error = error['source'][0]
                log.error(f"Error detected for ID {orig_id} in {tablename}",error=error)
        else:
            log.error(f"Collection {orig_id} in {tablename} not created for unknown reasons", error=res)

def produce_excel_template(config, null_itemname_only, batch_size, output, input_log):
    setup_globals(config, input_log, {'conn', 'log'})

    locked = Protection(locked=True, hidden=False)

    header_style = NamedStyle(name='header')
    header_style.font = Font(bold=True)
    header_style.protection = locked

    desc_style = NamedStyle(name='desc')
    desc_style.protection = locked
    side = Side(style="thick", color="000000")
    desc_style.border = Border(bottom=side)

    locked_style = NamedStyle(name='locked')
    locked_style.protection = locked

    styles = {header_style, desc_style, locked_style}

    headers = {
        'itemid': 'item identifier',
        'itemname': 'name field, used as ASpace title by default',
        'itemdesc': 'description of item',
        'title': 'title field, NOT ASpace title',
        'collid': 'db id of collection',
        'col##': 'human readable collection id',
        'collection title': 'title of collection',
        'title_override': 'Edit this to supply item title',
        'collection_override': 'Edit this to associate record with a collection whose title/EADID will be set to this'
    }

    sql_base = dedent(f"""\
    SELECT itemid, itemname, itemdesc, title, {tablename}.collid, "coll##", "collection title"
    FROM tblitems
    LEFT JOIN {resource_tablename} ON {tablename}.collid = {resource_tablename}.collid
    WHERE deptcodeid = {dept_id}
    AND itemid !~* '^[L|R]'
    {'AND itemname IS NULL' if null_itemname_only else ''}
    ORDER BY {tablename}.collid, itemid""")
    with conn.cursor() as cur:
        if batch_size:
            cur.arraysize = batch_size
        cur.execute(sql_base)
        method_to_use = 'fetchmany' if batch_size else 'fetchall'

        # Filename handling for output
        suffix = 0
        parts = [output, '', ".xlsx"] # initial xlsx has no suffix

        while rows := getattr(cur, method_to_use)():
            wb = openpyxl.Workbook()
            for style in styles:
                wb.add_named_style(style)
            ws = wb.active
            ws.append(list(headers.keys()))
            ws.append(list(headers.values()))
            for cell in ws[1]:
                cell.style = 'header'
            for cell in ws[2]:
                cell.style = 'desc'
            counter = 0
            for row in rows:
                counter += 1
                if counter % 100 == 0:
                    log.info('100 rows processed')
                ws.append(row)
                for cell in ws[ws.max_row]:
                    cell.style = 'locked'
            wb.save("".join(parts))
            log.info(f"Saved {''.join(parts)}")
            # Only relevant in batched mode
            suffix += 1
            parts[1] = f"_{suffix:02}"
